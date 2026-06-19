# # Отчет по 101 форме ЦБ
# ### Импорты
from pathlib import Path
import pandas as pd
from dbfread import DBF
import glob
import os
import re
import sys
import numpy as np
import py7zz
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.formatting.rule import ColorScaleRule
import requests
from datetime import datetime
from dateutil.relativedelta import relativedelta


# ### Скачивание архивов 101 формы и данных ОБС

save_dir = r"D:\Downloads" # <--------------------- Заменить на путь скачивания архивов
os.makedirs(save_dir, exist_ok=True)

def download(url, name):
    path = os.path.join(save_dir, name)
    try:
        r = requests.get(url, timeout=10)
        r.raise_for_status()
        with open(path, 'wb') as f:
            f.write(r.content)
        return True
    except:
        return False

last = max([re.search(r'(\d{8})', f).group(1) for f in os.listdir(save_dir) if re.match(r'101-\d{8}\.rar', f)])
new_date = (datetime.strptime(last, "%Y%m%d") + relativedelta(months=1)).strftime("%Y%m%d")

# Скачиваем
try:
    # Скачиваем 101 архив
    if not download(f"https://cbr.ru/vfs/credit/forms/101-{new_date}.rar", f"101-{new_date}.rar"):
        raise Exception("101 архив не скачался")


except Exception as e: # <--------------------- Вывод try-except при попытке скачивания
    print(f"\n❌ НОВЫЕ ДАННЫЕ ЕЩЕ НЕ ПОЯВИЛИСЬ")
    sys.exit(1)

download("https://cbr.ru/Content/Document/File/115862/obs_tabl20%D1%81.xlsx", "obs_tabl20с.xlsx")


# ### Распаковка архивов из загрузок в папку "101_rar"

# Находим все RAR файлы 101-* в загрузках
downloads = Path(save_dir) 
rar_files = list(downloads.glob("101-*.rar"))

if not rar_files:
    print("Архивы не найдены!")
    sys.exit(1)

# Распаковываем в папку 101_rar
extract_to = Path("101_rar") # <--------------------- В эту папку распакуются все архивы
extract_to.mkdir(exist_ok=True)

for rar in rar_files:
    try:
        # Распаковываем архив
        py7zz.extract_archive(str(rar), str(extract_to))
    except Exception as e:
        print(f"❌ Ошибка при распаковке {rar.name}: {e}")


# ### Обработка DBF
base_path = '101_rar'

# Находим все B1 файлы
b1_files = glob.glob(os.path.join(base_path, '*B1.dbf'))

# Читаем все B1 файлы
b1_dfs = []
for filepath in sorted(b1_files):
    filename = os.path.basename(filepath)

    # Извлекаем дату из имени файла (012026B1.dbf -> 01.2026)
    month = filename[:2]
    year = filename[2:6]
    file_date = pd.to_datetime(f"{year}-{month}-01")

    # Добавляем 1 месяц
    rep_dt = file_date + relativedelta(months=1)

    # Читаем файл
    table = DBF(filepath, encoding='cp866', ignore_missing_memofile=True)
    df = pd.DataFrame(table)

    # Добавляем столбец с датой
    df['rep_dt'] = rep_dt

    b1_dfs.append(df)

# Склеиваем все B1 файлы
if b1_dfs:
    df_b1_all = pd.concat(b1_dfs, ignore_index=True)

    # Удаляем дубли: оставляем строки с максимальной rep_dt
    df_b1_all = df_b1_all.sort_values('rep_dt').drop_duplicates(
        subset=['REGN', 'NUM_SC', 'DT', 'A_P'], 
        keep='last'
    )

# 2. СБОР ВСЕХ N1 ФАЙЛОВ с правильным декодированием
n1_files = []
pattern = os.path.join(base_path, '*N1.dbf')
n1_files.extend(glob.glob(pattern))

def decode_dbf_record(record):
    """Декодирует байтовые поля в строках записи"""
    decoded = {}
    for key, value in record.items():
        if isinstance(value, bytes):
            # Пробуем декодировать как CP866, удаляем нулевые байты в конце
            try:
                # Сначала удаляем нулевые байты и пробелы в конце
                clean_bytes = value.rstrip(b'\x00').strip()
                if clean_bytes:
                    decoded[key] = clean_bytes.decode('cp866')
                else:
                    decoded[key] = ''
            except:
                # Если не декодируется, оставляем как есть или преобразуем в int если это число
                try:
                    # Пробуем интерпретировать как число (для полей типа REGN)
                    if key == 'REGN' and len(value) <= 4:
                        # Для REGN может быть little-endian integer
                        import struct
                        if len(value) == 4:
                            num_val = struct.unpack('<I', value)[0]  # little-endian unsigned int
                            decoded[key] = str(num_val)
                        else:
                            decoded[key] = str(int.from_bytes(value, 'little'))
                    else:
                        decoded[key] = str(value)
                except:
                    decoded[key] = str(value)
        else:
            decoded[key] = value
    return decoded

# Читаем все N1 файлы с декодированием
n1_dfs = []
for filepath in sorted(n1_files):
    filename = os.path.basename(filepath)

    table = DBF(filepath, encoding='cp866', ignore_missing_memofile=True, raw=True)

    decoded_records = [decode_dbf_record(record) for record in table]
    df = pd.DataFrame(decoded_records)

    n1_dfs.append(df)

# Склеиваем все N1 файлы
if n1_dfs:
    df_n1_all = pd.concat(n1_dfs, ignore_index=True)

# 3. УДАЛЕНИЕ ДУБЛЕКАТОВ ПО REGN В N1
if not df_n1_all.empty and 'REGN' in df_n1_all.columns:
    # Приводим REGN к строковому типу для единообразия
    df_n1_all['REGN'] = df_n1_all['REGN'].astype(str)
    df_n1_all = df_n1_all.drop_duplicates(subset=['REGN'], keep='first')

# 4. MERGE N1 К B1 ПО ПОЛЮ REGN

# Преобразуем оба в строку
df_b1_all['REGN'] = df_b1_all['REGN'].astype(str)
df_n1_all['REGN'] = df_n1_all['REGN'].astype(str)

df_merged = df_b1_all.merge(df_n1_all, on='REGN', how='left', suffixes=('_b1', '_n1'))

#ПЕРЕВОДИМ СУММЫ ИЗ ТЫС. РУБ В РУБЛИ

df_merged['IITG'] = pd.to_numeric(df_merged['IITG'], errors='coerce')
# Умножаем на 1000 и создаем новую колонку
df_merged['IITG_RUB'] = df_merged['IITG'] * 1000

# Нужные колонки
cols = ['REGN', 'NAME_B', 'NUM_SC', 'DT', 'A_P', 'IITG_RUB']
df_merged = df_merged.drop_duplicates()
# Фильтруем по A_P == 1 и выбираем колонки
if 'A_P' in df_merged.columns:
    df_merged['A_P'] = df_merged['A_P'].astype(str).str.strip()
    df_final = df_merged[df_merged['A_P'] == '1'][cols].copy()

# Подтягиваем признак из Счета.xlsx

df_acc = pd.read_excel('Счета.xlsx') # <--------------------- Файл "Счета.xlsx" нужен для фильтрации по счетам, должен быть в папке
df_acc['key'] = df_acc['Счет'].astype(str).str[:3]

# Берем уникальные соответствия
mapping = dict(zip(df_acc['key'], df_acc['Применяется']))

# Применяем к нашему файлу
df_final['Применяется'] = df_final['NUM_SC'].astype(str).str[:3].map(mapping)

output_folder = Path(r"D:\Downloads\Исходные данные 101") # <------------ Сохранение сырых данных для нас в отдельную папку
output_folder.mkdir(parents=True, exist_ok=True)
df_merged.to_csv(output_folder / f"Исходные_данные_{new_date}.csv", index=False, encoding='utf-8-sig')
df_final.to_csv(output_folder / f"Активные_счета_{new_date}.csv", index=False, encoding='utf-8-sig')


# ## Данные ОБС
# Читаем файл, 4-я строка как шапка, пропускаем первые 3 строки
obs = pd.read_excel(
    r"D:\Downloads\obs_tabl20с.xlsx", # <--------------------- Заменить на путь файла "Динамические ряды ..."
    sheet_name="Активы - всего",
    header=3  # 4-я строка (0-индексация)
)

# Удаляем столбец А
obs = obs.iloc[:, 1:]

# Оставляем только нужные строки
need_codes = ['7.1.1.1.1', '7.1.1.1.3', '7.1.1.1.4', '7.1.1.2.1', 
              '7.1.1.2.3', '7.1.1.2.4', '7.1.1.3.1', '7.1.1.3.3', '7.1.2.3']
obs_filtered = obs[obs["№ п.п."].astype(str).isin(need_codes)]

obs_melted = pd.melt(
    obs_filtered,
    id_vars=['№ п.п.', ' Показатель (млрд руб.)'],  
    var_name='Дата',  
    value_name='Сумма ОБС'  
)

# Преобразуем даты в правильный формат
obs_melted['Дата'] = pd.to_datetime(obs_melted['Дата'], format='%d.%m.%y')
obs_melted = obs_melted.sort_values(['№ п.п.', 'Дата'])
obs_melted = obs_melted.reset_index(drop=True)

# фильтрация (Заменить на вводные данные)
obs_melted = obs_melted[obs_melted['Дата'] >= '2025-01-01']

# ## СТРОИМ ТАБЛИЦЫ

# ### СУММА РЫНКА ПО ДАННЫМ ОБС
df1 = obs_melted[obs_melted['№ п.п.'] != '7.1.2.3']
market = df1.groupby('Дата').agg({'Сумма ОБС': 'sum'}).reset_index()
market['Рынок_ОБС_руб'] = market['Сумма ОБС'] * 1e9
market = market.drop(columns='Сумма ОБС')

# ### ДОЛЯ ПРОСРОЧКИ
prosr = obs_melted[obs_melted["№ п.п."].astype(str).isin(['7.1.1.1.4', '7.1.1.2.4', '7.1.1.3.3', '7.1.2.3'])]
dole_df = (prosr.groupby("Дата").apply(
    lambda x: x[x["№ п.п."].astype(str) == "7.1.2.3"]["Сумма ОБС"].sum() / x["Сумма ОБС"].sum()
).reset_index(name="Доля_проср_физиков"))

# ### ДАННЫЕ 101 С КОРРЕКТИРОВКОЙ
df_final = df_final[
    (df_final['Применяется'] == 'Да') | 
    (df_final['NUM_SC'].astype(str).str.strip().isin(['47.1', '45.0']))
]
df_101 = df_final.copy()

df_final['DT'] = pd.to_datetime(df_101['DT']).dt.strftime('%Y-%m-%d')
df_101['DT'] = pd.to_datetime(df_101['DT']).dt.strftime('%Y-%m-%d')
dole_df['Дата'] = pd.to_datetime(dole_df['Дата']).dt.strftime('%Y-%m-%d')

# Словарь долей
dole_dict = dole_df.set_index('Дата')['Доля_проср_физиков'].to_dict()
mask = (df_101['NUM_SC'] == "458")
df_101['Корректировка'] = df_101['IITG_RUB'].astype(float)
df_101.loc[mask, 'Корректировка'] = df_101.loc[mask, 'IITG_RUB'].astype(float) * (1 - df_101.loc[mask, 'DT'].map(dole_dict).fillna(0))

# Итоговые таблицы
df_final = df_final[['REGN', 'NAME_B', 'NUM_SC', 'A_P', 'Применяется', 'DT', 'IITG_RUB']]
df_final.columns = ['Рег.н.', 'Наименование банка', 'Код счета', 'Тип счета', 'Применяется', 'Дата', 'Исх. остаток (руб.)']
result = df_101[['REGN', 'NAME_B', 'NUM_SC', 'A_P', 'Применяется', 'DT', 'IITG_RUB', 'Корректировка']]
result.columns = ['Рег.н.', 'Наименование банка', 'Код счета', 'Тип счета', 'Применяется', 'Дата', 'Исх. остаток (руб.)', 'Исх. остаток - корректировка (руб.)']

# ### РЫНОК ПО ДАННЫМ 101 БЕЗ КОРРЕКТИРОВКИ
res_market_2 = result.groupby('Дата').agg({'Исх. остаток (руб.)': 'sum'}).reset_index()
market['Дата'] = market['Дата'].astype(str)
res_market_2['Дата'] = res_market_2['Дата'].astype(str)
res_market_2['Исх. остаток (руб.)'] = res_market_2['Исх. остаток (руб.)'].astype(float)
both_market_2 = market[['Дата', 'Рынок_ОБС_руб']].merge(res_market_2, on='Дата', how='inner')
both_market_2['"Серая зона без корректировки" (ОБС-101)'] = both_market_2['Рынок_ОБС_руб'] - both_market_2['Исх. остаток (руб.)']
grouped_result_2 = result.groupby(['Рег.н.', 'Наименование банка', 'Дата'], as_index=False)['Исх. остаток (руб.)'].sum()

grey_zone_2 = both_market_2[['Дата', '"Серая зона без корректировки" (ОБС-101)']].copy()
grey_zone_2.columns = ['Дата', 'Исх. остаток (руб.)']
grey_zone_2['Рег.н.'] = '—'
grey_zone_2['Наименование банка'] = 'Серая зона (без корректировки)'
grouped_with_grey_2 = pd.concat([grouped_result_2, grey_zone_2], ignore_index=True)
grouped_with_grey_2 = grouped_with_grey_2.sort_values(['Дата', 'Рег.н.']).reset_index(drop=True)

# ### РЫНОК ПО ДАННЫМ 101 С КОРРЕКТИРОВКОЙ
res_market = result.groupby('Дата').agg({'Исх. остаток - корректировка (руб.)': 'sum'}).reset_index()
market['Дата'] = market['Дата'].astype(str)
res_market['Дата'] = res_market['Дата'].astype(str)
both_market = market.merge(res_market, on='Дата', how='inner')
both_market['"Серая зона" (ОБС-101)'] = both_market['Рынок_ОБС_руб'] - both_market['Исх. остаток - корректировка (руб.)']

grouped_result = result.groupby(['Рег.н.', 'Наименование банка', 'Дата'], as_index=False)['Исх. остаток - корректировка (руб.)'].sum()
grey_zone = both_market[['Дата', '"Серая зона" (ОБС-101)']].copy()
grey_zone.columns = ['Дата', 'Исх. остаток - корректировка (руб.)']
grey_zone.insert(0, 'Рег.н.', '—')  
grey_zone.insert(1, 'Наименование банка', 'Серая зона')
grouped_with_grey = pd.concat([grouped_result, grey_zone], ignore_index=True)
grouped_with_grey = grouped_with_grey.sort_values(['Дата', 'Рег.н.']).reset_index(drop=True)

# ## +Свод по Топ-банкам
obs_date = datetime.strptime(obs_melted['Дата'].max().strftime("%Y%m%d"), "%Y%m%d")
new_date_dt = datetime.strptime(new_date, "%Y%m%d")

# Флаг: доступны ли данные за новый период в xlsx
xlsx_available = obs_date and obs_date >= new_date_dt

# ### Без корректировки

# Преобразуем и фильтруем
grouped_with_grey_2['Дата'] = pd.to_datetime(grouped_with_grey_2['Дата'])
grouped_with_grey_2 = grouped_with_grey_2[grouped_with_grey_2['Дата'] >= grouped_with_grey_2['Дата'].max() - pd.DateOffset(months=3)]

# Возвращаем формат как было (YYYY-MM-DD)
grouped_with_grey_2['Дата'] = grouped_with_grey_2['Дата'].dt.strftime('%Y-%m-%d')

# Список топ банков
top_banks = [
    'СБЕРБАНК РОССИИ',
    'ВТБ',
    'ГАЗПРОМБАНК',
    'АЛЬФА-БАНК',
    'РОССЕЛЬХОЗБАНК',
    'МОСКОВСКИЙ КРЕДИТНЫЙ БАНК',
    'БАНК ДОМ.РФ',
    'СОВКОМБАНК',
    'ТБанк',
    'РАЙФФАЙЗЕНБАНК',
    'ЮНИКРЕДИТ БАНК'
]

# Создаем сводную таблицу
pivot = grouped_with_grey_2.pivot_table(
    index='Наименование банка',
    columns='Дата',
    values='Исх. остаток (руб.)',
    aggfunc='first'
).reset_index()

pivot.columns.name = None

# Делим все числовые столбцы на 1e9
numeric_cols = pivot.select_dtypes(include=['number']).columns
pivot[numeric_cols] = pivot[numeric_cols] / 1e9

# Фильтруем топ банки и сортируем в нужном порядке
df_top = pivot[pivot['Наименование банка'].isin(top_banks)]
df_top['Наименование банка'] = pd.Categorical(df_top['Наименование банка'], categories=top_banks, ordered=True)
df_top = df_top.sort_values('Наименование банка').reset_index(drop=True)

grey_zone = pivot[pivot['Наименование банка'] == 'Серая зона (без корректировки)']

# Прочие банки (все, кроме топов и серой зоны)
others = pivot[~pivot['Наименование банка'].isin(top_banks + ['Серая зона (без корректировки)'])]
others_sum = others.drop('Наименование банка', axis=1).sum().to_frame().T
others_sum.insert(0, 'Наименование банка', 'ПРОЧИЕ')

# Объединяем
result = pd.concat([df_top, grey_zone, others_sum], ignore_index=True)

# Добавляем строку итогов по топ банкам
top_sum = df_top.drop('Наименование банка', axis=1).sum().to_frame().T
top_sum.insert(0, 'Наименование банка', 'Общий итог ТОП банков')

# Добавляем строку общего итога
total_sum = pivot.drop('Наименование банка', axis=1).sum().to_frame().T
total_sum.insert(0, 'Наименование банка', 'Общий итог')

# Финальная таблица
final = pd.concat([result, top_sum, total_sum], ignore_index=True)


growth = final.copy()

date_cols = [col for col in growth.columns if col != 'Наименование банка']

# Рассчитываем приросты для каждого банка
for i in range(1, len(date_cols)):
    growth[date_cols[i]] = final[date_cols[i]] - final[date_cols[i-1]]

# Удаляем первый месяц
growth = growth.drop(columns=[date_cols[0]])

relative = final.copy()

date_cols = [col for col in relative.columns if col != 'Наименование банка']

# Рассчитываем относительные приросты для каждого банка
for i in range(1, len(date_cols)):
    # Избегаем деления на ноль
    denominator = final[date_cols[i-1]]
    relative[date_cols[i]] = (final[date_cols[i]] - denominator) / denominator.where(denominator != 0, np.nan)

# Первый месяц удаляем
relative = relative.drop(columns=[date_cols[0]])

# Заменяем бесконечности на NaN
relative = relative.replace([np.inf, -np.inf], np.nan)

final.drop(final.columns[1], axis=1, inplace=True)


# ### С корректировкой

# Преобразуем и фильтруем
grouped_with_grey['Дата'] = pd.to_datetime(grouped_with_grey['Дата'])
grouped_with_grey = grouped_with_grey[grouped_with_grey['Дата'] >= grouped_with_grey['Дата'].max() - pd.DateOffset(months=3)]
grouped_with_grey['Дата'] = grouped_with_grey['Дата'].dt.strftime('%Y-%m-%d')

if xlsx_available:
    # Список топ банков
    top_banks = [
        'СБЕРБАНК РОССИИ',
        'ВТБ',
        'ГАЗПРОМБАНК',
        'АЛЬФА-БАНК',
        'РОССЕЛЬХОЗБАНК',
        'МОСКОВСКИЙ КРЕДИТНЫЙ БАНК',
        'БАНК ДОМ.РФ',
        'СОВКОМБАНК',
        'ТБанк',
        'РАЙФФАЙЗЕНБАНК',
        'ЮНИКРЕДИТ БАНК'
    ]

    # Создаем сводную таблицу
    pivot = grouped_with_grey.pivot_table(
        index='Наименование банка',
        columns='Дата',
        values='Исх. остаток - корректировка (руб.)',
        aggfunc='first'
    ).reset_index()

    pivot.columns.name = None

    # Делим все числовые столбцы на 1e9
    numeric_cols = pivot.select_dtypes(include=['number']).columns
    pivot[numeric_cols] = pivot[numeric_cols] / 1e9

    # Фильтруем топ банки и сортируем в нужном порядке
    df_top = pivot[pivot['Наименование банка'].isin(top_banks)]
    df_top['Наименование банка'] = pd.Categorical(df_top['Наименование банка'], categories=top_banks, ordered=True)
    df_top = df_top.sort_values('Наименование банка').reset_index(drop=True)

    grey_zone = pivot[pivot['Наименование банка'] == 'Серая зона']

    # Прочие банки (все, кроме топов и серой зоны)
    others = pivot[~pivot['Наименование банка'].isin(top_banks + ['Серая зона'])]
    others_sum = others.drop('Наименование банка', axis=1).sum().to_frame().T
    others_sum.insert(0, 'Наименование банка', 'ПРОЧИЕ')

    # Объединяем
    result = pd.concat([df_top, grey_zone, others_sum], ignore_index=True)

    # Добавляем строку итогов по топ банкам
    top_sum = df_top.drop('Наименование банка', axis=1).sum().to_frame().T
    top_sum.insert(0, 'Наименование банка', 'Общий итог ТОП банков')

    # Добавляем строку общего итога
    total_sum = pivot.drop('Наименование банка', axis=1).sum().to_frame().T
    total_sum.insert(0, 'Наименование банка', 'Общий итог')

    # Финальная таблица
    final_cor = pd.concat([result, top_sum, total_sum], ignore_index=True)

    growth_cor = final_cor.copy()

    date_cols = [col for col in growth_cor.columns if col != 'Наименование банка']

    # Рассчитываем приросты для каждого банка (используем growth, а не final)
    for i in range(1, len(date_cols)):
        growth_cor[date_cols[i]] = final_cor[date_cols[i]] - final_cor[date_cols[i-1]]

    # Удаляем первый месяц
    growth_cor = growth_cor.drop(columns=[date_cols[0]])

    relative_cor = final_cor.copy()

    date_cols = [col for col in relative_cor.columns if col != 'Наименование банка']

    # Рассчитываем относительные приросты для каждого банка
    for i in range(1, len(date_cols)):
        # Избегаем деления на ноль
        denominator = final_cor[date_cols[i-1]]
        relative_cor[date_cols[i]] = (final_cor[date_cols[i]] - denominator) / denominator.where(denominator != 0, np.nan)

    # Первый месяц удаляем
    relative_cor = relative_cor.drop(columns=[date_cols[0]])

    # Заменяем бесконечности на NaN
    relative_cor = relative_cor.replace([np.inf, -np.inf], np.nan)

    final_cor.drop(final_cor.columns[1], axis=1, inplace=True)

else:
    final_cor = pd.DataFrame({'Сообщение': ['Актуальные данные с корректировкой еще не появились']})
    growth_cor = pd.DataFrame({'Сообщение': ['Актуальные данные с корректировкой еще не появились']})
    relative_cor = pd.DataFrame({'Сообщение': ['Актуальные данные с корректировкой еще не появились']})


# ### Сохраненяем

with pd.ExcelWriter(f'Сводные с приростами_{new_date}.xlsx', engine='openpyxl') as writer: # <---------- Сохранение сводных таблиц

    # Получаем книгу
    workbook = writer.book

    # === СОЗДАЕМ ЛИСТЫ ===
    # Переименовываем стандартный лист
    if 'Sheet' in workbook.sheetnames:
        workbook['Sheet'].title = 'Сводные без корректировки'
    else:
        workbook.create_sheet('Сводные без корректировки')

    # Создаем второй лист
    if 'Сводные с корректировкой' not in workbook.sheetnames:
        workbook.create_sheet('Сводные с корректировкой')

    # === ЛИСТ 1: Без корректировки ===
    start_row = 0
    ws = workbook['Сводные без корректировки']

    # Таблица 1
    ws.cell(row=start_row + 1, column=1, value='Абсолютные значения, млрд. руб.')
    final.to_excel(writer, sheet_name='Сводные без корректировки', startrow=start_row + 1, index=False)
    start_row += len(final) + 4

    # Таблица 2
    ws.cell(row=start_row + 1, column=1, value='Абсолютные приросты, млрд. руб.')
    growth.to_excel(writer, sheet_name='Сводные без корректировки', startrow=start_row + 1, index=False)
    start_row += len(growth) + 4

    # Таблица 3
    ws.cell(row=start_row + 1, column=1, value='Относительные приросты, %')
    relative.to_excel(writer, sheet_name='Сводные без корректировки', startrow=start_row + 1, index=False)

    # === ЛИСТ 2: С корректировкой ===
    start_row = 0
    ws = workbook['Сводные с корректировкой']

    # Таблица 1
    ws.cell(row=start_row + 1, column=1, value='Абсолютные значения (с корректировкой), млрд. руб.')
    final_cor.to_excel(writer, sheet_name='Сводные с корректировкой', startrow=start_row + 1, index=False)
    start_row += len(final_cor) + 4

    # Таблица 2
    ws.cell(row=start_row + 1, column=1, value='Абсолютные приросты (с корректировкой), млрд. руб.')
    growth_cor.to_excel(writer, sheet_name='Сводные с корректировкой', startrow=start_row + 1, index=False)
    start_row += len(growth_cor) + 4

    # Таблица 3
    ws.cell(row=start_row + 1, column=1, value='Относительные приросты (с корректировкой), %')
    relative_cor.to_excel(writer, sheet_name='Сводные с корректировкой', startrow=start_row + 1, index=False)

    # === Форматирование ===
    from openpyxl.styles import Font
    from openpyxl.formatting.rule import ColorScaleRule

    for sheet_name in ['Сводные без корректировки', 'Сводные с корректировкой']:
        ws = workbook[sheet_name]

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and ('Абсолютные' in cell.value or 'Относительные' in cell.value):
                    table_start_row = cell.row + 1
                    table_end_row = table_start_row + 16

                    cell.font = Font(bold=True, size=12)

                    for col in range(1, ws.max_column + 1):
                        header_cell = ws.cell(row=table_start_row, column=col)
                        if header_cell.value:
                            header_cell.font = Font(bold=True)

                    for row_idx in range(table_start_row + 1, table_end_row + 1):
                        if ws.cell(row=row_idx, column=1).value and 'Общий итог' in str(ws.cell(row=row_idx, column=1).value):
                            for col in range(1, ws.max_column + 1):
                                ws.cell(row=row_idx, column=col).font = Font(bold=True)

                    is_percent = 'Относительные' in cell.value
                    for row_idx in range(table_start_row + 1, table_end_row + 1):
                        for col in range(2, ws.max_column + 1):
                            cell_data = ws.cell(row=row_idx, column=col)
                            if isinstance(cell_data.value, (int, float)):
                                if is_percent:
                                    cell_data.number_format = '0.00%'
                                else:
                                    cell_data.number_format = '#,##0.00'

                    if is_percent:
                        if table_end_row >= table_start_row + 1 and ws.max_column >= 2:
                            start_cell = ws.cell(row=table_start_row + 1, column=2).coordinate
                            end_cell = ws.cell(row=table_end_row, column=ws.max_column).coordinate
                            color_scale = ColorScaleRule(
                                start_type='percentile', start_value=5, start_color='D73027',
                                mid_type='percentile', mid_value=50, mid_color='FFFFFF',
                                end_type='percentile', end_value=95, end_color='1A9641'
                            )
                            ws.conditional_formatting.add(f'{start_cell}:{end_cell}', color_scale)

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width

